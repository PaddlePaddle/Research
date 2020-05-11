#-*- coding:utf-8 -*-
# zhangwei99@baidu.com
import openpyxl as pyxl
import numpy as np
import math
import os
from collections import OrderedDict


def parseXLSX(xlsxName, parseArray):
    wb = pyxl.load_workbook(xlsxName)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    for i in range(2, ws.max_row+1):
        cvalue = ws.cell(row = i, column = 2).value
        if cvalue is not None:
            fId = int(cvalue)
        else:
            import pdb; pdb.set_trace()
        cvalue = ws.cell(row = i, column = 3).value
        if cvalue is not None:
            mId = int(cvalue)
        else:
            import pdb; pdb.set_trace()
        cvalue = ws.cell(row = i, column = 4).value
        if cvalue is not None:
            if cvalue == "car" or cvalue == "car ":
                type = 0 
            elif cvalue == "truck" or cvalue == "truch":
                type = 1 
            else:
                import pdb; pdb.set_trace()
    
        parseArray[fId, mId - 1, type] = 1


#def parseTxt(TxtName, parseArray):
#    fp = open(TxtName)
#    lines = fp.readlines()
#    for line in lines:
#        sp = line.rstrip("\n").split(",")
#        fId = int(sp[1])
#        if fId >= 3000:
#            continue
#        mId = int(sp[2])
#        type = int(sp[3])
#        parseArray[fId, mId - 1, type] = 1


def parseTxt(TxtName, parseArray):
    fp = open(TxtName)
    lines = fp.readlines()
    for line in lines:
        sp = line.rstrip("\n").split(" ")
        fId = int(sp[1]) - 1 
        if fId >= 3000:
            continue
        mId = int(sp[2])
        type = int(sp[3]) - 1
        parseArray[fId, mId - 1, type] = 1


def compute_nwRMSE(n, pdArray, gtArray):
    # weight
    wVect = np.asarray(np.arange(1, n+1)) / (n * (n + 1) / 2.0)
    fNum, mNum, typeNum = pdArray.shape
    lst = range(0, fNum)
    interval = int(math.ceil(fNum / float(n)))
    segLsts = [lst[i : i + interval] for i in range(0, len(lst), interval)]
    gtCntArray = np.zeros(mNum)
    pdCntArray = np.zeros(mNum)
    nwRMSEArray = np.zeros((mNum, 2))
    wRMSEArray = np.zeros((mNum, 2))
    vehicleNumArray = np.zeros((mNum, 2))
    for mId in range(0, mNum):
        gtCntArray[mId] = np.sum(gtArray[:, mId, :])
        pdCntArray[mId] = np.sum(pdArray[:, mId, :])
        for tId in range(0, 2):
            # wRMSE
            diffVectCul = np.zeros(n)
            for segId, frames in enumerate(segLsts):
                diff = np.square(sum(pdArray[0:frames[-1], mId, tId]) - sum(gtArray[0:frames[-1], mId, tId]))
                diffVectCul[segId] = diff
            wRMSE = np.sqrt(np.dot(wVect, diffVectCul))

            # num
            vehicleNum = np.sum(gtArray[:, mId, tId])
            vehicleNumArray[mId, tId] = vehicleNum 

            # for print only
            if vehicleNum == 0:
                wRMSEArray[mId, tId] = 0 
            else:
                wRMSEArray[mId, tId] = wRMSE / vehicleNum

            #nwRMSE
            if wRMSE > vehicleNum:
                nwRMSE = 0
            else:
                if vehicleNum == 0:
                    nwRMSE = 0
                else:
                    nwRMSE = 1 - wRMSE / vehicleNum
            nwRMSEArray[mId, tId] = nwRMSE

    print("")
    printStr = " moveID: "
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "% 4d | "%(moveId+1)
    print(printStr)
    printStr = " --------"
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "-------"
    print(printStr)

    printStr = " gt cnt: "
    for val in gtCntArray.tolist():
        printStr += "%04d | "%(val)
    print(printStr)
    printStr = " --------"
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "-------"
    print(printStr)

    printStr = " pd cnt: "
    for val in pdCntArray.tolist():
        printStr += "%04d | "%(val)
    print(printStr)
    printStr = " --------"
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "-------"
    print(printStr)

    printStr = " nwRMSE: "
    for moveId, val in enumerate(np.sum(nwRMSEArray, axis=1).tolist()):
        printStr += "%2.2f | "%(val)
    print(printStr)

    nwRMSEArray = np.multiply(nwRMSEArray, vehicleNumArray)
    return np.sum(nwRMSEArray), np.sum(vehicleNumArray) 


if __name__ == "__main__":
    videoInfo = {"cam_1":{"frame_num":3000, "movement_num":4},
                 "cam_1_dawn":{"frame_num":3000, "movement_num":4},
                 "cam_1_rain":{"frame_num":2961, "movement_num":4},
                 "cam_2":{"frame_num":18000, "movement_num":4},
                 "cam_2_rain":{"frame_num":3000, "movement_num":4},
                 "cam_3":{"frame_num":18000, "movement_num":4},
                 "cam_3_rain":{"frame_num":3000, "movement_num":4},
                 "cam_4":{"frame_num":27000, "movement_num":12},
                 "cam_4_dawn":{"frame_num":4500, "movement_num":12},
                 "cam_4_rain":{"frame_num":3000, "movement_num":12},
                 "cam_5":{"frame_num":18000, "movement_num":12},
                 "cam_5_dawn":{"frame_num":3000, "movement_num":12},
                 "cam_5_rain":{"frame_num":3000, "movement_num":12},
                 "cam_6":{"frame_num":18000, "movement_num":12},
                 "cam_6_snow":{"frame_num":3000, "movement_num":12},
                 "cam_7":{"frame_num":14400, "movement_num":12},
                 "cam_7_dawn":{"frame_num":2400, "movement_num":12},
                 "cam_7_rain":{"frame_num":3000, "movement_num":12},
                 "cam_8":{"frame_num":3000, "movement_num":6},
                 "cam_9":{"frame_num":3000, "movement_num":12},
                 "cam_10":{"frame_num":2111, "movement_num":3},
                 "cam_11":{"frame_num":2111, "movement_num":3},
                 "cam_12":{"frame_num":1997, "movement_num":3},
                 "cam_13":{"frame_num":1966, "movement_num":3},
                 "cam_14":{"frame_num":3000, "movement_num":2},
                 "cam_15":{"frame_num":3000, "movement_num":2},
                 "cam_16":{"frame_num":3000, "movement_num":2},
                 "cam_17":{"frame_num":3000, "movement_num":2},
                 "cam_18":{"frame_num":3000, "movement_num":2},
                 "cam_19":{"frame_num":3000, "movement_num":2},
                 "cam_20":{"frame_num":3000, "movement_num":2}}

    # segment number
    n = 10 

    gtXlsxRoot = "./gt/"
    pdXlsxRoot = "./vehicle_counting_results/"
    vNum = len(videoInfo.keys())
    nwRMSEVec = np.zeros(vNum)
    vehicleNumVec = np.zeros(vNum) 
    vId = 0
    for vName, info in videoInfo.items():
        fNum = videoInfo[vName]["frame_num"]
        if fNum > 3000:
            fNum = 3000
        mNum = videoInfo[vName]["movement_num"]

        # parse gt
        gtArray = np.zeros((fNum, mNum, 2))
        gtXlsx = gtXlsxRoot + "/" + vName + ".xlsx"
        if not os.path.exists(gtXlsx):
            continue
        parseXLSX(gtXlsx, gtArray)

        # parse prediction 
        pdArray = np.zeros((fNum, mNum, 2))
        pdXlsx = pdXlsxRoot + "/" + vName + ".txt"
        if not os.path.exists(pdXlsx):
            continue
        parseTxt(pdXlsx, pdArray)

        nwRMSE, vehicleNum = compute_nwRMSE(n, pdArray, gtArray)
        nwRMSEVec[vId] = nwRMSE
        vehicleNumVec[vId] = vehicleNum
        vId += 1
        print(" %s nwRMSE: %f"%(vName, nwRMSE/vehicleNum))

    score2 = sum(nwRMSEVec) / sum(vehicleNumVec)

    baseFactor = 0.464906
    videoTotal = 300 + 296 + 300 + 300 + 30 * 60 + 300 + 30 * 60 + 300 + 300 + 30 * 60 + 300 + 300 + 30 * 60 + 300 + 30 * 60 + 300 + 300 + 30 * 60 + 300 + 300 + 211 + 211 + 200 + 197 + 300 + 300 + 300 + 300 + 300 + 300 + 300
    #time = 6217 
    time = 9997 # res50
    time = 11418 # res50 pipeline
    #time = 43642 # res154
    #time = 8487 # omni 
    score1 = 1 - (time * baseFactor) / (5 * float(videoTotal)) 

    score = 0.3 * score1 + 0.7 * score2
    print("\ns1: %f; effective: %f; efficient: %f"%(score, score2, score1))
